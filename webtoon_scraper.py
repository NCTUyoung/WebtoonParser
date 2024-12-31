from web_scraper_template import WebScraper
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import logging
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 設置日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename=f'scraping_{datetime.now().strftime("%Y%m%d")}.log'
)

class WebtoonScraper(WebScraper):
    def __init__(self, base_url, title_no):
        self.base_url = base_url
        self.title_no = title_no
        super().__init__(self.get_url(1))  # 初始化第一頁
        self.headers.update({
            'Referer': 'https://www.webtoons.com',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7'
        })

    def get_url(self, page):
        """生成特定頁碼的URL"""
        return f"{self.base_url}?title_no={self.title_no}&page={page}"

    def get_total_pages(self, soup):
        """獲取總頁數"""
        try:
            # 找到頁碼列表
            pagination = soup.find('div', class_='paginate')
            if pagination:
                # 獲取最後一個頁碼
                pages = pagination.find_all('a')
                if pages:
                    return int(pages[-1].text)
            return 1
        except Exception as e:
            logging.error(f"獲取總頁數時發生錯誤: {e}")
            return 1

    def scrape_all_pages(self):
        """爬取所有頁面的內容"""
        all_data = {
            '章節標題': [],
            '發布日期': [],
            '點讚數': [],
            '章節編號': []
        }

        # 獲取第一頁內容來確定總頁數
        html_content = self.get_page_content()
        if not html_content:
            return None

        soup = self.parse_content(html_content)
        if not soup:
            return None

        total_pages = self.get_total_pages(soup)
        logging.info(f"總頁數: {total_pages}")

        # 爬取每一頁
        for page in range(1, total_pages + 1):
            logging.info(f"正在爬取第 {page} 頁")
            
            if page > 1:
                self.url = self.get_url(page)
                html_content = self.get_page_content()
                if not html_content:
                    continue
                soup = self.parse_content(html_content)
                if not soup:
                    continue

            page_data = self.parse_webtoon_content(soup)
            if page_data:
                for key in all_data:
                    all_data[key].extend(page_data[key])

            # 添加延遲避免請求過快
            time.sleep(2)

        return all_data

    def parse_webtoon_content(self, soup):
        """解析WEBTOON頁面內容"""
        try:
            data = {
                '章節標題': [],
                '發布日期': [],
                '點讚數': [],
                '章節編號': []
            }

            episodes = soup.find_all('li', class_='_episodeItem')
            
            for episode in episodes:
                title = episode.find('span', class_='subj').text.strip()
                date = episode.find('span', class_='date').text.strip()
                likes = episode.find('span', class_='like_area').text.strip()
                episode_num = episode.find('span', class_='tx').text.strip()

                data['章節標題'].append(title)
                data['發布日期'].append(date)
                data['點讚數'].append(likes)
                data['章節編號'].append(episode_num)

            return data
            
        except Exception as e:
            logging.error(f"解析WEBTOON內容時發生錯誤: {e}")
            return None

    def get_webtoon_info(self, soup):
        """抓取漫畫基本信息"""
        try:
            info = {
                '標題': '',
                '觀看數': '',
                '訂閱數': '',
                '評分': '',
                '更新時間': '',
                '描述': '',
                '作者': '',
                '抓取時間': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }

            # 抓取標題
            title = soup.find('h1', class_='subj')
            if title:
                info['標題'] = title.text.strip()

            # 抓取觀看數
            view = soup.find('span', class_='ico_view')
            if view and view.find_next('em'):
                info['觀看數'] = view.find_next('em').text.strip()

            # 抓取訂閱數
            subscribe = soup.find('span', class_='ico_subscribe')
            if subscribe and subscribe.find_next('em'):
                info['訂閱數'] = subscribe.find_next('em').text.strip()

            # 抓取評分
            grade = soup.find('em', class_='cnt')
            if grade:
                info['評分'] = grade.text.strip()

            # 修改更新时间抓取
            update_info = soup.find('p', class_='day_info')  # 更新class名称
            if update_info:
                info['更新時間'] = update_info.text.strip()
            else:
                # 备选方案：尝试找含有特定文本的span元素
                update_span = soup.find('span', class_='date')
                if update_span:
                    info['更新時間'] = update_span.text.strip()

            # 抓取描述
            description = soup.find('p', class_='summary')
            if description:
                info['描述'] = description.text.strip()

            # 抓取作者
            author = soup.find('h2', text='作家')
            if author and author.find_next():
                info['作者'] = author.find_next().text.strip()

            return info

        except Exception as e:
            logging.error(f"抓取漫畫信息時發生錯誤: {e}")
            return None

    def save_webtoon_info(self, info, filename):
        """保存漫畫基本信息到CSV"""
        try:
            # 將字典轉換為單行DataFrame
            df = pd.DataFrame([info])
            df.to_csv(filename, index=False, encoding='utf-8-sig')
            logging.info(f"漫畫基本信息已保存到 {filename}")
        except Exception as e:
            logging.error(f"保存漫畫信息時發生錯誤: {e}")

    def save_to_excel(self, webtoon_info, chapters_data, filename):
        """將漫畫信息和章節數據保存到Excel文件"""
        try:
            # 創建工作簿
            wb = Workbook()
            
            # 創建漫畫信息工作表
            info_ws = wb.active
            info_ws.title = "漫畫資訊"
            
            # 將漫畫信息轉換為DataFrame並寫入
            info_df = pd.DataFrame([webtoon_info])
            for r_idx, row in enumerate(dataframe_to_rows(info_df, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    info_ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 創建章節數據工作表
            chapters_ws = wb.create_sheet(title="章節列表")
            
            # 將章節數據轉換為DataFrame並寫入
            chapters_df = pd.DataFrame(chapters_data)
            for r_idx, row in enumerate(dataframe_to_rows(chapters_df, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    chapters_ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 調整欄寬
            for ws in [info_ws, chapters_ws]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存Excel文件
            wb.save(filename)
            logging.info(f"數據已保存到Excel文件: {filename}")
            
        except Exception as e:
            logging.error(f"保存Excel文件時發生錯誤: {e}")

def main():
    base_url = "https://www.webtoons.com/zh-hant/bl-gl/friday-night/list"
    title_no = "6875"
    
    scraper = WebtoonScraper(base_url, title_no)
    
    webtoon_info = None
    chapters_data = None
    
    # 獲取第一頁內容來抓取基本信息
    html_content = scraper.get_page_content()
    if html_content:
        soup = scraper.parse_content(html_content)
        if soup:
            # 抓取漫畫基本信息
            webtoon_info = scraper.get_webtoon_info(soup)
    
    # 抓取所有章節信息
    chapters_data = scraper.scrape_all_pages()
    
    # 如果兩種數據都成功獲取，則保存到Excel
    if webtoon_info and chapters_data:
        excel_filename = f'webtoon_data_{datetime.now().strftime("%Y%m%d")}.xlsx'
        scraper.save_to_excel(webtoon_info, chapters_data, excel_filename)
        logging.info(f"成功將所有數據保存到Excel文件")
    else:
        logging.error("無法獲取完整數據")

if __name__ == "__main__":
    main() 